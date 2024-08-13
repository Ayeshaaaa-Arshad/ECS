import openpyxl
from django.core.management.base import BaseCommand
from django.db import IntegrityError
from django.contrib.auth.models import Group
from users.models import Customer, AppAdmin, ShopAdmin, User

class Command(BaseCommand):
    help = 'Insertion or updating of Customers, ShopAdmins, and AppAdmins from an Excel file based on ID'

    def add_arguments(self, parser):
        parser.add_argument(
            '--mode',
            choices=['insert', 'update'],
            required=True,
            help='Specify the mode of operation: insert or update'
        )

    def handle(self, *args, **options):
        mode = options['mode']
        if mode not in ['insert', 'update']:
            self.stdout.write(self.style.ERROR('Invalid mode selected. Choose between "insert" and "update".'))
            return

        # Loading the workbook
        workbook = openpyxl.load_workbook('users/Users.xlsx')

        errors = []

        # Data lists for bulk operations
        customers_to_create = []
        app_admins_to_create = []
        shop_admins_to_create = []
        users_to_update = []

        # Fetch groups
        customer_group, _ = Group.objects.get_or_create(name='Customer')
        app_admin_group, _ = Group.objects.get_or_create(name='AppAdmin')
        shop_admin_group, _ = Group.objects.get_or_create(name='ShopAdmin')

        # Process Customer sheet
        customer_sheet = workbook['Customers']
        for row in customer_sheet.iter_rows(min_row=2, values_only=True):
            user_id, first_name, last_name, username, email, phone_number, address, group_name, password = row

            phone_number = str(phone_number).strip()

            if len(phone_number) > 11:
                phone_number = phone_number[:11]

            try:
                if mode == 'insert':
                    if User.objects.filter(id=user_id).exists():
                        errors.append(f"Error: Customer ID {user_id} already exists.")
                        continue

                    user = User(
                        id=user_id,
                        first_name=first_name,
                        last_name=last_name,
                        username=username,
                        email=email,
                        phone_number=phone_number,
                        address=address
                    )
                    user.set_password(password)
                    user.save()
                    user.groups.add(customer_group)
                    customers_to_create.append(Customer(user=user))
                elif mode == 'update':
                    user, created = User.objects.update_or_create(
                        id=user_id,
                        defaults={
                            'first_name': first_name,
                            'last_name': last_name,
                            'username': username,
                            'email': email,
                            'phone_number': phone_number,
                            'address': address
                        }
                    )
                    user.set_password(password)
                    user.groups.add(customer_group)
                    if created:
                        customers_to_create.append(Customer(user=user))
                    else:
                        users_to_update.append(user)

            except IntegrityError as e:
                errors.append(f'Error processing Customer ID {user_id}: {str(e)}')

        # Process AppAdmin sheet
        app_admin_sheet = workbook['AppAdmins']
        for row in app_admin_sheet.iter_rows(min_row=2, values_only=True):
            user_id, first_name, last_name, username, email, phone_number, address, group_name,password = row

            phone_number = str(phone_number).strip()

            if len(phone_number) > 11:
                phone_number = phone_number[:11]

            try:
                if mode == 'insert':
                    if User.objects.filter(email=email).exists():
                        errors.append(f"Error: AppAdmin ID {user_id} with {username} already exists.")
                        continue

                    user = User(
                        id=user_id,
                        first_name=first_name,
                        last_name=last_name,
                        username=username,
                        email=email,
                        phone_number=phone_number,
                        address=address
                    )
                    user.set_password(password)
                    user.save()
                    user.groups.add(app_admin_group)
                    app_admins_to_create.append(AppAdmin(user=user))
                elif mode == 'update':
                    user, created = User.objects.update_or_create(
                        id=user_id,
                        defaults={
                            'first_name': first_name,
                            'last_name': last_name,
                            'username': username,
                            'email': email,
                            'phone_number': phone_number,
                            'address': address
                        }
                    )
                    user.set_password(password)

                    user.groups.add(app_admin_group)
                    if created:
                        app_admins_to_create.append(AppAdmin(user=user))
                    else:
                        users_to_update.append(user)

            except IntegrityError as e:
                errors.append(f'Error processing AppAdmin ID {user_id}: {str(e)}')

        # Process ShopAdmin sheet
        shop_admin_sheet = workbook['ShopAdmins']
        for row in shop_admin_sheet.iter_rows(min_row=2, values_only=True):
            user_id, first_name, last_name, username, email, phone_number, address, group_name,password = row

            phone_number = str(phone_number).strip()

            if len(phone_number) > 11:
                phone_number = phone_number[:11]

            try:
                if mode == 'insert':
                    if User.objects.filter(id=user_id).exists():
                        errors.append(f"Error: ShopAdmin ID {user_id} already exists.")
                        continue

                    user = User(
                        id=user_id,
                        first_name=first_name,
                        last_name=last_name,
                        username=username,
                        email=email,
                        phone_number=phone_number,
                        address=address
                    )
                    user.set_password(password)
                    user.save()
                    user.groups.add(shop_admin_group)
                    shop_admins_to_create.append(ShopAdmin(user=user))
                elif mode == 'update':
                    user, created = User.objects.update_or_create(
                        id=user_id,
                        defaults={
                            'first_name': first_name,
                            'last_name': last_name,
                            'username': username,
                            'email': email,
                            'phone_number': phone_number,
                            'address': address
                        }
                    )
                    user.set_password(password)
                    print(user.set_password(password))
                    user.groups.add(shop_admin_group)
                    if created:
                        shop_admins_to_create.append(ShopAdmin(user=user))
                    else:
                        users_to_update.append(user)

            except IntegrityError as e:
                errors.append(f'Error processing ShopAdmin ID {user_id}: {str(e)}')

        # Bulk operations
        if mode == 'insert':
            Customer.objects.bulk_create(customers_to_create)
            AppAdmin.objects.bulk_create(app_admins_to_create)
            ShopAdmin.objects.bulk_create(shop_admins_to_create)
        elif mode == 'update':
            User.objects.bulk_update(users_to_update, ['first_name', 'last_name', 'email', 'phone_number', 'address'])

        # Display all errors at the end
        if errors:
            self.stdout.write(self.style.ERROR('The following errors occurred:'))
            for error in errors:
                self.stdout.write(self.style.ERROR(error))

        self.stdout.write(self.style.SUCCESS('Data processing complete'))
